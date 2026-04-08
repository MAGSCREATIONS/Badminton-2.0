import io
import os

from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side
from logic import (
    load_players, add_player, remove_player, generate_balanced_groups,
    mark_attendance,
    get_attendance_by_date, get_all_attendance, get_attendance_by_player,
)
from session_db import (
    create_session, get_session, get_all_sessions,
    record_payment, get_player_summary, clear_all_payments
)

app = Flask(__name__)

@app.route("/")
def home():
    return render_template("index.html", players=load_players())

@app.route("/add_player", methods=["POST"])
def add():
    data = request.json
    add_player(data['name'], data['skill'])
    return jsonify({"success": True})

@app.route("/remove_player", methods=["POST"])
def remove():
    data = request.json
    remove_player(data['name'])
    return jsonify({"success": True})

@app.route("/randomize", methods=["POST"])
def randomize():
    data = request.json
    present_names = data.get("present_players", [])
    try:
        groups = generate_balanced_groups(present_names)
        output = [
            {
                "group_number": g.get_group_number(),
                "team1": [p.get_name() for p in g.get_team1().get_players()],
                "team2": [p.get_name() for p in g.get_team2().get_players()]
            }
            for g in groups
        ]
        return jsonify({"groups": output})
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/mark_attendance", methods=["POST"])
def attendance_mark():
    data = request.json
    present_names = data.get("present_players", [])
    date = data.get("date", None)
    try:
        summary = mark_attendance(present_names, date=date)
        return jsonify(summary)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/attendance/today")
def attendance_today():
    return jsonify(get_attendance_by_date())

@app.route("/attendance/<date>")
def attendance_by_date(date):
    try:
        return jsonify(get_attendance_by_date(date))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

@app.route("/attendance/player/<player_name>")
def attendance_by_player(player_name):
    return jsonify(get_attendance_by_player(player_name))

@app.route("/attendance/all")
def attendance_all():
    return jsonify(get_all_attendance())

# ── Session routes ────────────────────────────────────────────

@app.route("/session/create", methods=["POST"])
def session_create():
    data = request.json
    try:
        session = create_session(
            date=data.get("date"),
            courts=int(data["courts"]),
            session_hours=float(data["hours"]),
            player_hours=data["player_hours"]
        )
        return jsonify(session)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/session/<date>")
def session_get(date):
    session = get_session(date)
    if not session:
        return jsonify({"error": "No session found"}), 404
    return jsonify(session)

@app.route("/session/all")
def session_all():
    return jsonify(get_all_sessions())

@app.route("/session/pay", methods=["POST"])
def session_pay():
    data = request.json
    try:
        updated = record_payment(data["date"], data["player_name"], float(data["amount"]))
        return jsonify(updated)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@app.route("/session/summary")
def session_summary():
    return jsonify(get_player_summary())

@app.route("/session/clear-payments", methods=["POST"])
def session_clear_payments():
    try:
        return jsonify(clear_all_payments())
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# ── Excel Export Feature (ADDED) ───────────────────────────────

def create_excel_from_sessions():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    right_align = Alignment(horizontal="right", vertical="center")

    thin = Side(style='thin')
    thick = Side(style='medium')

    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_left_thick = Border(left=thick, right=thin, top=thin, bottom=thin)
    border_right_thick = Border(left=thin, right=thick, top=thin, bottom=thin)

    yellow_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    sessions = get_all_sessions()
    row_num = 1

    for session in sessions:
        for p in session.get("players", []):
            val1 = p.get("amount_due", 0)
            val2 = p.get("amount_paid", 0)
            diff = val2 - val1

            c1 = ws.cell(row=row_num, column=1, value=val1)
            c2 = ws.cell(row=row_num, column=2, value=val2)
            c3 = ws.cell(row=row_num, column=3, value=diff)

            for c in (c1, c2, c3):
                c.number_format = '#,##0.00'
                c.alignment = right_align

            # Borders
            c1.border = border_left_thick
            c2.border = border
            c3.border = border_right_thick

            # Coloring (matches screenshot)
            if diff == 0:
                c3.fill = orange_fill
            elif diff < 0:
                c3.fill = yellow_fill

            row_num += 1

    # Column widths (tight layout)
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return stream


@app.route("/session/export_excel")
def export_excel():
    file_stream = create_excel_from_sessions()
    return send_file(
        file_stream,
        as_attachment=True,
        download_name="session_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ── Run ───────────────────────────────────────────────────────

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
